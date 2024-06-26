import os
import time
from fastapi import FastAPI, File, UploadFile, HTTPException,Response
import io
import PyPDF2
import re
from pypdf import PdfReader 
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from xlsxwriter import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
app = FastAPI()

def read_pdf_text(file):
    pdf_reader = PyPDF2.PdfFileReader(file,strict=False)
    text = ""
    for page_num in range(pdf_reader.numPages):
        page = pdf_reader.getPage(page_num)
        text += page.extractText()
    return text
def create_excel_with_data(file_path, data):
    # Create a new Workbook
    wb = Workbook()

    # Select the active worksheet
    ws = wb.active
    ws.title="Translate"

    # Add data to the worksheet
    for row in data:
        # ws.cell (row=1,column=row+1).value = ILLEGAL_CHARACTERS_RE.sub(r'',row)
        ws.append(row)

    # Save the workbook
    wb.save(file_path)
    return

# Example usage



@app.post("/pdf/text/")
async def read_pdf_text_endpoint(file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Uploaded file is not a PDF.")
    
    pdf_content = await file.read()
    pdf_file = io.BytesIO(pdf_content)
    text = read_pdf_text(pdf_file)
    result = re.split(r'\d+', text)
    data=[]
    result = [item for item in result if item]
    # print(result)
    driver = webdriver.Chrome()
    driver.get("https://www.google.com/search?q=english+to+tamil")
    text_input = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//textarea[@placeholder='Enter text']"))
        )
    text_output = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='tw-target-text-container']//pre/span[1]"))
        )
    data = [
    ['English', 'Tamil'],]
    for val in result:
         text_input.send_keys(val)
         time.sleep(5)
         print(text_output.text) 
         data2=[val,text_output.text]
         data.append(data2)
         text_input.clear()
         print(data)
    # print(data)    
    driver.close()
    create_excel_with_data("file/Translated.xlsx", data) 
    output = io.BytesIO()
    
    output.seek(0)
    file_path="file/Translated.xlsx"

    # Set response headers for file download
   
    
    # dls = "Translated.xlsx"
    # urllib.request.urlretrieve(dls, "test.xls") 
    # Check if the file exists
    if not os.path.exists(file_path):
        return Response(content="File not found", status_code=404)
    
    # Set content disposition to 'attachment' to trigger a file download
    headers = {
        "Content-Disposition": f"attachment; filename={os.path.basename(file_path)}"
    }
    
    # Return a response with the file as content and headers set
    return Response(content=open(file_path, "rb").read(), headers=headers)
# @app.post("/upload2")
# async def upload2(file: UploadFile = File(...)):


  
# # creating a pdf reader object 
#     reader = PdfReader(file) 
  
# # printing number of pages in pdf file 
#     print(len(reader.pages)) 
  
# # getting a specific page from the pdf file 
#     page = reader.pages[0] 
  
# # extracting text from page 
#     text = page.extract_text() 
#     print(text) 
#     return text
@app.get("/trastlate")
def transtlate(file_path: str, headers: dict, items: list):
    
    with Workbook(file_path) as workbook:
        worksheet = workbook.add_worksheet()
        worksheet.write_row(row=0, col=0, data=headers.values())
        header_keys = list(headers.keys())
        for index, item in enumerate(items):
            row = map(lambda field_id: item.get(field_id, ''), header_keys)
            worksheet.write_row(row=index + 1, col=0, data=row)

headers = {
    'id': 'User Id',
    'name': 'Full Name',
    'rating': 'Rating',
}

items = [
    {'id': 1, 'name': "Ilir Meta", 'rating': 0.06},
    {'id': 2, 'name': "Abdelmadjid Tebboune", 'rating': 4.0},
    {'id': 3, 'name': "Alexander Lukashenko", 'rating': 3.1},
    {'id': 4, 'name': "Miguel Díaz-Canel", 'rating': 0.32}
]

transtlate("my-xlsx-file.xlsx", headers, items)
@app.post("/englishtotamil/file")
async def englishtotamil(file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Uploaded file is not a PDF.")
    
    pdf_content = await file.read()
    pdf_file = io.BytesIO(pdf_content)
    text = read_pdf_text(pdf_file)
    result = re.split(r'\d+', text)
    headers = {
    'id':"Id",
    'english': 'English',
    'tamil': 'Tamil',
    'v3rdPersonSingular':'V 3rd Person Singular',
    'present':'Present',
    'presentContinue':'Present Continue',
    'past':'Past',
    'future':'Future',
}
    
    result = [item for item in result if item]
    # print(result)
    driver = webdriver.Chrome()
    driver.get("https://www.google.com/search?q=english+to+tamil")
    driver.execute_script("window.open('https://www.wordreference.com/es/translation.asp?tranword=')")
    text_input = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//textarea[@placeholder='Enter text']"))
        )
    text_output = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='tw-target-text-container']//pre/span[1]"))
        )
  
    
    windows = driver.window_handles
    id=0
    items = []
    for val in result:
        driver.switch_to.window(windows[0])
        text_input.clear()
        if(val=="APEUniPTEVocabList" or val=="Visitwww.apeuni.comformorestudymaterials" or val=="APEUniPTEVocabListVisitwww.apeuni.comformorestudymaterialsAPEUniPTEBasicVocab"):
            continue
        print(val)
        if(val.lower().endswith('Pageof')):
            val.replace('Pageof', '')
        text_input.send_keys(val)
        time.sleep(2)
        translated=text_output.text
        driver.switch_to.window(windows[1])
        time.sleep(1)
        
        driver.get("https://www.wordreference.com/es/translation.asp?tranword="+val)
        # verb_input = WebDriverWait(driver, 10).until(
        #     EC.element_to_be_clickable((By.XPATH, "//input[@type='search']"))
        # )
        try:
           v3rdPersonSingular = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='inflectionsSection']//div/dl/dd[1]/dl/dt")) )
           v3rdPersonSingular.text
        except:
            
            v3rdPersonSingular="--"

        try:
           presentContinue = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@class='inflectionsSection']//div/dl/dd[2]/dl/dt"))
        )
           presentContinue.text
        except:
           presentContinue="--"
        try:
           past = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@class='inflectionsSection']//div/dl/dd[3]/dl/dt"))
        )
           past.text
        except:
           past="--"
        try:
            present = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@class='inflectionsSection']//div/dl/dd[4]/dl/dt"))
        )
           present="Wil "+present.text
        except:
           present="--" 
        id +=1
       
        # print(id,translated+"==",v3rdPersonSingular+"==",present+"==",presentContinue+"==",past+"==",present+"==")
        item1={"id":id,"english":val,"tamil":translated,"v3rdPersonSingular":v3rdPersonSingular,"present":present,"presentContinue":presentContinue,"past":past,"future":present}
       
        items.append(item1)
        
        print(items)
    # print(data)    
    driver.close()
    fileName1:str=file.filename.replace(".pdf","")
    print(fileName1)
    with Workbook(fileName1+".xlsx") as workbook:
        worksheet = workbook.add_worksheet()
        worksheet.write_row(row=0, col=0, data=headers.values())
        header_keys = list(headers.keys())
        for index, item in enumerate(items):
            row = map(lambda field_id: item.get(field_id, ''), header_keys)
            worksheet.write_row(row=index + 1, col=0, data=row)
    return {"message":"Done"}